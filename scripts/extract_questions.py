"""One-time script: parse the xlsx spreadsheet into questions.json and pre-compute embeddings."""

import json
import sys
from pathlib import Path

import numpy as np
import openpyxl
from sentence_transformers import SentenceTransformer

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = BASE_DIR / "OneDrive_1_2-27-2026" / "paul_jennings_all_questions - updated.xlsx"
DATA_DIR = BASE_DIR / "data"
VIDEO_DIR = BASE_DIR / "OneDrive_1_2-27-2026"

# Column C files with "(1)" in the filename
C_RENAMED = {5, 6, 8, 9}


def video_path(col: str, row: int) -> str | None:
    """Return the relative video path for a given column+row, or None if missing."""
    if row == 37 and col in ("E", "F", "G", "H"):
        folder = "Extra Videos"
        filename = f"{col}37.mp4"
    elif col in ("B", "C", "D"):
        folder = f"Column {col}"
        if col == "C" and row in C_RENAMED:
            filename = f"{col}{row} (1).mp4"
        else:
            filename = f"{col}{row}.mp4"
    else:
        return None

    full = VIDEO_DIR / folder / filename
    if not full.exists():
        print(f"  WARNING: video not found: {full}")
        return None
    return f"{folder}/{filename}"


def main():
    DATA_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Trimmed"]

    questions = []

    for row_idx in range(2, 38):  # rows 2-37
        question_text = ws.cell(row=row_idx, column=1).value or ""
        answer_b = ws.cell(row=row_idx, column=2).value or ""
        answer_c = ws.cell(row=row_idx, column=3).value or ""
        answer_d = ws.cell(row=row_idx, column=4).value or ""
        notes = ws.cell(row=row_idx, column=5).value or ""

        entry = {
            "id": row_idx - 1,  # 1-36, 37=fallback
            "row": row_idx,
            "question": question_text.strip(),
            "answers": {
                "B": answer_b.strip(),
                "C": answer_c.strip(),
                "D": answer_d.strip(),
            },
            "videos": {
                "B": video_path("B", row_idx),
                "C": video_path("C", row_idx),
                "D": video_path("D", row_idx),
            },
        }

        if notes.strip():
            entry["notes"] = notes.strip()

        # Row 37 fallback has extra variations in E-H
        if row_idx == 37:
            entry["is_fallback"] = True
            for col_letter, col_idx in [("E", 5), ("F", 6), ("G", 7), ("H", 8)]:
                answer = ws.cell(row=row_idx, column=col_idx).value or ""
                if answer.strip():
                    entry["answers"][col_letter] = answer.strip()
                    vp = video_path(col_letter, row_idx)
                    if vp:
                        entry["videos"][col_letter] = vp

        questions.append(entry)

    # Write questions.json
    out_path = DATA_DIR / "questions.json"
    with open(out_path, "w") as f:
        json.dump(questions, f, indent=2, ensure_ascii=False)
    print(f"Wrote {len(questions)} entries to {out_path}")

    # Pre-compute embeddings for the 36 real questions (not fallback)
    real_questions = [q["question"] for q in questions if not q.get("is_fallback")]
    print(f"Computing embeddings for {len(real_questions)} questions...")
    model = SentenceTransformer("all-MiniLM-L6-v2")
    embeddings = model.encode(real_questions, normalize_embeddings=True)

    emb_path = DATA_DIR / "embeddings.npy"
    np.save(emb_path, embeddings)
    print(f"Saved embeddings ({embeddings.shape}) to {emb_path}")

    # Verify
    print("\nQuestions extracted:")
    for q in questions:
        tag = " [FALLBACK]" if q.get("is_fallback") else ""
        vids = sum(1 for v in q["videos"].values() if v)
        print(f"  {q['id']:2d}. {q['question'][:60]}{tag} ({vids} videos)")


if __name__ == "__main__":
    main()
